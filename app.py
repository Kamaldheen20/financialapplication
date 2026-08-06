# ==========================
# IMPORTS
# ==========================
from datetime import datetime
from sqlalchemy.exc import IntegrityError
import os
import secrets
import logging
from dotenv import load_dotenv
import io

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    session
)


from flask_login import (
    LoginManager,
    login_user,
    login_required,
    logout_user,
    current_user
)

from werkzeug.security import (
    generate_password_hash,
    check_password_hash
)

from openpyxl import Workbook
from models import (
    db,
    Admin,
    Customer,
    Payment
)

# ==========================
# APP CONFIG
# ==========================

import sys

# ---------------------------------------------------------------------
# FIX: locate .env correctly whether running as a normal script or as
# a PyInstaller-built .exe.
#
# - Normal `python app.py`: base dir = folder containing this file.
# - Frozen .exe (PyInstaller): sys.frozen is True and sys.executable
#   points at the .exe itself. We use the folder the .exe lives in
#   (NOT sys._MEIPASS, which is a temp extraction folder that gets
#   deleted after the app closes and isn't where you'd want users
#   editing credentials anyway).
#
# This means: ship a `.env` file in the SAME FOLDER as the built
# .exe (e.g. dist/YourApp/.env), and users can edit DB credentials
# there without rebuilding the exe.
# ---------------------------------------------------------------------
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

env_path = os.path.join(BASE_DIR, ".env")
load_dotenv(env_path)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

app = Flask(__name__)

# ---------------------------------------------------------------------
# FIX (Security): the old fallback was a fixed, publicly-known string
# ("change-this-secret-key"). If SECRET_KEY was ever left unset, every
# session cookie in production would be signed with a secret an
# attacker could read straight out of this source file, letting them
# forge login sessions. We now fall back to a random key instead, and
# log a loud warning so the missing .env value gets noticed. Sessions
# just won't survive an app restart until SECRET_KEY is set properly -
# no worse than before, but no longer a static, guessable secret.
# ---------------------------------------------------------------------
_secret_key = os.getenv("SECRET_KEY")
if not _secret_key:
    _secret_key = secrets.token_hex(32)
    logger.warning(
        "SECRET_KEY not set in .env - using a random key for this run. "
        "Sessions will be invalidated on every restart. Set SECRET_KEY in "
        "your .env file for stable, secure sessions."
    )
app.config["SECRET_KEY"] = _secret_key

# Harden session cookies for production (safe no-ops for local/dev use).
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
# NOTE: defaults to False because the PyInstaller desktop build serves
# the UI over plain http://127.0.0.1 (no TLS) - a Secure cookie would
# silently break login there. Set SESSION_COOKIE_SECURE=true in the
# Render .env (HTTPS-only deployment) to enable it there.
app.config["SESSION_COOKIE_SECURE"] = os.getenv("SESSION_COOKIE_SECURE", "false").lower() == "true"
# Cap upload size (restore_database accepts file uploads) to prevent
# large-file / decompression-bomb style denial-of-service uploads.
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024  # 25 MB

database_url = os.getenv("DATABASE_URL")

if not database_url:
    # Fail loudly with a clear, actionable message instead of letting
    # Flask-SQLAlchemy raise its generic "SQLALCHEMY_DATABASE_URI must
    # be set" RuntimeError with no context about WHY it's missing.
    error_msg = (
        f"DATABASE_URL not found.\n\n"
        f"Expected a .env file at:\n{env_path}\n\n"
        f"containing a line like:\n"
        f"DATABASE_URL=postgresql://user:password@host:5432/dbname\n\n"
        f"Place a .env file next to the .exe and restart the app."
    )
    if getattr(sys, "frozen", False):
        # Show a native Windows message box since there's no console
        # attached to a windowed PyInstaller build to print to.
        import ctypes
        ctypes.windll.user32.MessageBoxW(0, error_msg, "Configuration Error", 0x10)
        sys.exit(1)
    else:
        raise RuntimeError(error_msg)

app.config["SQLALCHEMY_DATABASE_URI"] = database_url

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db.init_app(app)

login_manager = LoginManager()

login_manager.init_app(app)

login_manager.login_view = "login"


@login_manager.user_loader
def load_user(user_id):
    # FIX: a tampered/stale session cookie could contain a non-numeric
    # user_id. int() on that previously raised an uncaught ValueError,
    # turning EVERY request (any page load) into a 500 error until the
    # cookie was manually cleared. Fail safe -> treat as logged-out.
    try:
        return db.session.get(Admin, int(user_id))
    except (TypeError, ValueError):
        return None

# ==========================
# COMPANY SETTINGS MODEL
# ==========================

class CompanySettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, nullable=False)
    company_name = db.Column(db.String(200))
    address = db.Column(db.String(500))
    phone = db.Column(db.String(50))


# ==========================
# PDF FONT HELPER
# Supports Tamil AND English (Latin) in the same PDF — including mixed
# names like "குmaரெசன்" (Tamil + Latin letters in one string).
#
# Root cause of the bug:
#   NotoSans-Regular does NOT contain Tamil glyphs (U+0B80–U+0BFF).
#   NotoSansTamil does NOT contain Latin glyphs.
#   Neither font alone can render mixed-script text.
#
# Fix — dual-font split-span approach:
#   1. Register NotoSansTamil  (for Tamil characters)
#   2. Register NotoSans       (for Latin/English characters)
#   3. _pdf_text(s) wraps every character in the correct <font> tag
#      so ReportLab Paragraph can render both scripts in one cell/line.
# ==========================

_FONT_CACHE = {}


def _find_font_file(candidates):
    for path in candidates:
        if path and os.path.exists(path):
            return path
    return None


def _download_font(url, save_path):
    import urllib.request
    try:
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        print(f"[PDF font] Downloading {os.path.basename(save_path)} …")
        urllib.request.urlretrieve(url, save_path)
        print(f"[PDF font] Saved to {save_path}")
        return True
    except Exception as e:
        print(f"[PDF font] Download failed: {e}")
        return False


def _ensure_font(fname, github_subpath):
    """Locate font on disk across common paths, or auto-download it."""
    app_fonts = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
    windir = os.environ.get("WINDIR", "C:\\Windows")

    candidates = [
        os.path.join(app_fonts, fname),
        os.path.join("C:\\FinanceManager", "fonts", fname),
        os.path.join(windir, "Fonts", fname),
        f"/usr/share/fonts/truetype/noto/{fname}",
        f"/usr/share/fonts/opentype/noto/{fname}",
        f"/Library/Fonts/{fname}",
        f"/System/Library/Fonts/{fname}",
    ]
    path = _find_font_file(candidates)
    if path:
        return path

    save_path = os.path.join(app_fonts, fname)
    url = (
        "https://github.com/googlefonts/noto-fonts/raw/main/"
        f"hinted/ttf/{github_subpath}"
    )
    return save_path if _download_font(url, save_path) else None


def _try_register(reg_name, path):
    """Register a TTFont. Returns True if already registered or success."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    try:
        pdfmetrics.getFont(reg_name)
        return True
    except KeyError:
        pass
    if path and os.path.exists(path):
        try:
            pdfmetrics.registerFont(TTFont(reg_name, path))
            return True
        except Exception as e:
            print(f"[PDF font] Cannot register {reg_name}: {e}")
    return False


def _register_pdf_fonts():
    """
    Register both NotoSansTamil and NotoSans with ReportLab.
    Returns (tamil_font, latin_font, bold_font, has_tamil).
    Always call this before building any PDF.
    """
    if "result" in _FONT_CACHE:
        return _FONT_CACHE["result"]

    # ── Tamil font (NotoSansTamil covers U+0B80–U+0BFF) ────────────────────
    tamil_path = _ensure_font(
        "NotoSansTamil-Regular.ttf",
        "NotoSansTamil/NotoSansTamil-Regular.ttf"
    )
    tamil_ok = _try_register("NotoSansTamil", tamil_path)
    if tamil_ok:
        _try_register("NotoSansTamil-Bold", tamil_path)  # reuse same file

    # ── Latin font (NotoSans covers A–Z, a–z, digits, punctuation) ─────────
    latin_path = _ensure_font(
        "NotoSans-Regular.ttf",
        "NotoSans/NotoSans-Regular.ttf"
    )
    latin_bold_path = _ensure_font(
        "NotoSans-Bold.ttf",
        "NotoSans/NotoSans-Bold.ttf"
    )
    latin_ok = _try_register("NotoSans", latin_path)
    if latin_ok:
        if not _try_register("NotoSans-Bold", latin_bold_path or latin_path):
            pass  # bold falls back to regular below

    if tamil_ok and latin_ok:
        print("[PDF font] Dual-font mode: NotoSansTamil + NotoSans (Tamil + Latin)")
        result = ("NotoSansTamil", "NotoSans", "NotoSans-Bold", True)
    elif tamil_ok:
        print("[PDF font] Tamil-only mode: NotoSansTamil (Latin may look basic)")
        result = ("NotoSansTamil", "NotoSansTamil", "NotoSansTamil-Bold", True)
    elif latin_ok:
        print("[PDF font] Latin-only mode: NotoSans (Tamil will not render)")
        result = ("NotoSans", "NotoSans", "NotoSans-Bold", False)
    else:
        print("[PDF font] WARNING: No Unicode fonts found. Using Helvetica.")
        result = ("Helvetica", "Helvetica", "Helvetica-Bold", False)

    _FONT_CACHE["result"] = result
    return result


def _pdf_text(text):
    """
    Wrap each character in the correct <font> tag so ReportLab Paragraph
    renders both Tamil and Latin characters correctly in the same string.

    Tamil Unicode block: U+0B80 – U+0BFF
    Everything else (Latin, digits, spaces, punctuation) uses the Latin font.
    """
    cached = _FONT_CACHE.get("result")
    if cached is None:
        _register_pdf_fonts()
        cached = _FONT_CACHE["result"]

    tamil_font = cached[0]
    latin_font = cached[1]

    # If no dual-font support, just return plain text
    if tamil_font == latin_font:
        return text

    result = []
    current_font = None
    for ch in str(text):
        cp = ord(ch)
        font = tamil_font if 0x0B80 <= cp <= 0x0BFF else latin_font
        if font != current_font:
            if current_font is not None:
                result.append("</font>")
            result.append(f'<font name="{font}">')
            current_font = font
        # Escape XML special characters
        if ch == "&":
            result.append("&amp;")
        elif ch == "<":
            result.append("&lt;")
        elif ch == ">":
            result.append("&gt;")
        else:
            result.append(ch)
    if current_font:
        result.append("</font>")
    return "".join(result)


# Backward-compat alias used throughout the rest of the file
def _register_tamil_font():
    tamil_font, latin_font, bold_font, has_tamil = _register_pdf_fonts()
    # Return (normal_font, bold_font, has_tamil) as before — callers use
    # the normal_font for body text; _pdf_text() handles per-char switching.
    return tamil_font, bold_font, has_tamil


# Pre-register at startup so the first PDF is fast
try:
    _register_pdf_fonts()
except Exception:
    pass


# ==========================
# CUSTOMER SORT HELPER
# Sorts customers by customer_id numerically (1, 2, 8, 22, 88, 888, 11112)
# instead of alphabetically (1, 11112, 2, 22, 8, 88, 888), since
# customer_id is stored as a string in the database.
# ==========================

def _sort_customers(customers):
    def _id_sort_key(c):
        try:
            return (0, int(c.customer_id))
        except (ValueError, TypeError):
            return (1, c.customer_id)
    return sorted(customers, key=_id_sort_key)


# ==========================
# CUSTOMER ALERT HELPER
# Used by the Customer Ledger page. Computes a finer-grained status than
# the plain Active/Closed `status` field:
#   - "Settled"             -> status is already Closed (fully paid off)
#   - "Collection Required" -> loan term (end_date) has passed but a
#                               balance is still owed
#   - "Active"               -> everything else (within term, still owing)
# NOTE: this is display-only. It's set as a plain Python attribute on
# each Customer object for template rendering and is never committed to
# the database.
# ==========================

def _compute_alert(customer):
    if customer.status == "Closed":
        return "Settled"

    try:
        if customer.end_date:
            end_dt = datetime.strptime(str(customer.end_date), "%Y-%m-%d")
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            if today > end_dt and (customer.remaining_balance or 0) > 0:
                return "Collection Required"
    except (ValueError, TypeError):
        pass  # malformed/missing end_date -> fall through to "Active"

    return "Active"


# ==========================
# HOME / REDIRECT
# ==========================

@app.route("/")
def home():
    return redirect(url_for("login"))


# ==========================
# LOGIN
# ==========================

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        admin = Admin.query.filter_by(username=username).first()

        if admin and check_password_hash(admin.password, password):
            login_user(admin)
            return redirect(url_for("dashboard"))

        flash("Invalid Username or Password")

    return render_template("login.html")


# ==========================
# REGISTER
# ==========================

@app.route("/register", methods=["GET", "POST"])
def register():
    

    if request.method == "POST":

        username = request.form["username"].strip()
        mobile = request.form["mobile"].strip()
        password = request.form["password"]

        # Check username
        if Admin.query.filter_by(username=username).first():
            flash("Username already exists", "danger")
            return redirect(url_for("register"))

        # Check mobile
        if Admin.query.filter_by(mobile=mobile).first():
            flash("Mobile number already registered", "danger")
            return redirect(url_for("register"))

        admin = Admin(
            username=username,
            mobile=mobile,
            password=generate_password_hash(password)
        )

        db.session.add(admin)
        db.session.commit()

        flash("Account Created Successfully", "success")
        return redirect(url_for("login"))
    return render_template("register.html")
#forget pass
@app.route("/forgot_password", methods=["GET", "POST"])
def forgot_password():

    if request.method == "POST":

        username = request.form["username"].strip()
        mobile = request.form["mobile"].strip()
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]

        # Check password confirmation
        if password != confirm_password:
            flash("Passwords do not match.", "danger")
            return redirect(url_for("forgot_password"))

        # Find the user
        admin = Admin.query.filter_by(
            username=username,
            mobile=mobile
        ).first()

        if not admin:
            flash(
                "Invalid Username or Mobile Number.",
                "danger"
            )
            return redirect(url_for("forgot_password"))

        # Update password
        admin.password = generate_password_hash(password)

        db.session.commit()

        flash(
            "Password updated successfully. Please login.",
            "success"
        )

        return redirect(url_for("login"))
    return render_template("forgot_password.html")
# ==========================
# DASHBOARD
# ==========================

@app.route("/dashboard")
@login_required
def dashboard():
    customers = Customer.query.filter_by(user_id=current_user.id).all()
    customers = _sort_customers(customers)

    total_customers = len(customers)
    active_customers = len([c for c in customers if c.status == "Active"])
    closed_customers = len([c for c in customers if c.status == "Closed"])

    total_loan = sum(c.loan_amount for c in customers)
    total_paid = sum(c.total_paid for c in customers)
    total_balance = sum(c.remaining_balance for c in customers)

    today = datetime.now().strftime("%Y-%m-%d")
    current_month = datetime.now().strftime("%Y-%m")

    # FIX (Performance): the old code loaded EVERY payment row for the
    # user into Python just to sum two numbers. At ~1,000 customers with
    # years of daily payment history that's an ever-growing, unbounded
    # query on every single dashboard load. Let Postgres do the sum.
    from sqlalchemy import func

    today_collection = db.session.query(
        func.coalesce(func.sum(Payment.amount), 0)
    ).filter(
        Payment.user_id == current_user.id,
        Payment.payment_date == today
    ).scalar()

    month_collection = db.session.query(
        func.coalesce(func.sum(Payment.amount), 0)
    ).filter(
        Payment.user_id == current_user.id,
        Payment.payment_date.like(f"{current_month}%")
    ).scalar()

    return render_template(
        "dashboard.html",
        customers=customers,
        total_customers=total_customers,
        active_customers=active_customers,
        closed_customers=closed_customers,
        total_loan=total_loan,
        total_paid=total_paid,
        total_balance=total_balance,
        today_collection=today_collection,
        month_collection=month_collection
    )


# ==========================
# ADD CUSTOMER
# ==========================
@app.route("/add_customer", methods=["POST"])
@login_required
def add_customer():

    customer_id = request.form["customer_id"].strip()

    existing = Customer.query.filter_by(
        customer_id=customer_id,
        user_id=current_user.id
    ).first()

    if existing:
        flash(f"Customer ID '{customer_id}' already exists.", "danger")
        return redirect(url_for("dashboard"))

    # FIX: float() on bad/missing input used to raise an uncaught
    # ValueError -> 500 error page instead of a friendly flash message.
    try:
        loan_amount = float(request.form["loan_amount"])
        daily_due = float(request.form["daily_due"])
    except (ValueError, TypeError):
        flash("Loan Amount and Daily Due must be valid numbers.", "danger")
        return redirect(url_for("dashboard"))

    if loan_amount < 0 or daily_due < 0:
        flash("Loan Amount and Daily Due cannot be negative.", "danger")
        return redirect(url_for("dashboard"))

    start_date = request.form.get("start_date", "")
    end_date = request.form.get("end_date", "")
    address = request.form.get("address", "").strip()

    if not end_date and start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            # Fixed 3-month loan term, regardless of loan amount / daily due
            month = start_dt.month - 1 + 3
            year = start_dt.year + month // 12
            month = month % 12 + 1
            day = min(start_dt.day, [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28,
                                      31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
            end_dt = start_dt.replace(year=year, month=month, day=day)
            end_date = end_dt.strftime("%Y-%m-%d")
        except ValueError:
            end_date = ""

    customer = Customer(
        customer_id=customer_id,
        name=request.form["name"],
        mobile=request.form["mobile"],
        address=address,
        loan_amount=loan_amount,
        daily_due=daily_due,
        total_paid=0,
        remaining_balance=loan_amount,
        status="Active",
        start_date=start_date,
        end_date=end_date,
        user_id=current_user.id
    )

    try:
        db.session.add(customer)
        db.session.commit()
        flash("Customer Added Successfully", "success")

    except IntegrityError:
        db.session.rollback()
        flash(f"Customer ID '{customer_id}' already exists.", "danger")

    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", "danger")

    return redirect(url_for("dashboard"))
# ==========================
# ADD PAYMENT
# ==========================

@app.route("/add_payment", methods=["POST"])
@login_required
def add_payment():
    # FIX: float() on bad/missing input used to raise an uncaught
    # ValueError -> 500 error page instead of a friendly flash message.
    try:
        amount = float(request.form["amount"])
    except (ValueError, TypeError):
        flash("Payment amount must be a valid number.", "danger")
        return redirect(url_for("dashboard"))

    if amount < 0:
        flash("Payment amount cannot be negative.", "danger")
        return redirect(url_for("dashboard"))

    # FIX: the dashboard's Daily Due Collection now uses a searchable
    # "Reg. No. - Name" field (datalist) instead of a <select>, backed by
    # a hidden #customer_id_hidden input that only gets filled in when the
    # typed text exactly matches a real customer. If nothing was matched
    # (blank search, JS didn't run, stale page, etc.) that hidden field
    # posts as "" - previously that empty string was accepted here, a
    # Payment row got created with customer_id="", and it silently never
    # showed up in the Collection Sheet / Daily Report / Customer Ledger
    # because nothing joins to an empty customer_id. Reject it up front
    # instead of saving an orphaned payment.
    customer_id = request.form.get("customer_id", "").strip()

    if not customer_id:
        flash("Please pick a customer from the search box before saving the payment.", "danger")
        return redirect(url_for("dashboard"))

    payment_date = request.form["payment_date"]

    # FIX (Transaction safety): lock the customer row for the duration of
    # this transaction with SELECT ... FOR UPDATE. Without this, two
    # payments submitted for the same customer at nearly the same time
    # could both read the same starting total_paid, and the second
    # commit would silently overwrite (lose) the first payment's balance
    # update - a real risk for a finance app where staff may double-tap
    # "Save" or two collectors record the same customer concurrently.
    # This is a no-op on SQLite but takes effect on PostgreSQL/Supabase.
    customer = Customer.query.filter_by(
        customer_id=customer_id,
        user_id=current_user.id
    ).with_for_update().first()

    # FIX: if the id made it through non-empty but still doesn't match any
    # customer of this user (typo, stale/duplicate id, wrong account),
    # the old code carried on anyway - inserting a Payment row that has
    # no matching Customer and, again, never appears in any report.
    if not customer:
        flash(f"No customer found with ID '{customer_id}'. Payment was not saved.", "danger")
        return redirect(url_for("dashboard"))

    # Check if a payment already exists for this customer on this date.
    # If it does, update it instead of inserting a duplicate row -
    # this prevents the amount from being double-counted when the
    # same id/date is submitted again (accidentally or otherwise).
    existing_payment = Payment.query.filter_by(
        customer_id=customer_id,
        payment_date=payment_date,
        user_id=current_user.id
    ).first()

    if existing_payment:
        old_amount = existing_payment.amount
        existing_payment.amount = amount
        amount_diff = amount - old_amount

        if customer:
            customer.total_paid += amount_diff

        flash("Existing payment for this date was updated (no duplicate created)")
    else:
        payment = Payment(
            customer_id=customer_id,
            payment_date=payment_date,
            amount=amount,
            user_id=current_user.id
        )
        db.session.add(payment)

        if customer:
            customer.total_paid += amount

        flash("Payment Added Successfully")

    if customer:
        customer.remaining_balance = customer.loan_amount - customer.total_paid

        if customer.remaining_balance <= 0:
            customer.remaining_balance = 0
            customer.status = "Closed"
        else:
            customer.status = "Active"

    db.session.commit()

    return redirect(url_for("dashboard"))


# ==========================
# COLLECTION SHEET
# ==========================

@app.route("/collection_sheet")
@login_required
def collection_sheet():
    selected_month = request.args.get(
        "month",
        datetime.now().strftime("%Y-%m")
    )

    customers = Customer.query.filter_by(user_id=current_user.id).all()
    payments = Payment.query.filter_by(user_id=current_user.id).all()

    customers = _sort_customers(customers)

    return render_template(
        "collection_sheet.html",
        customers=customers,
        payments=payments,
        selected_month=selected_month
    )


# ==========================
# DAILY REPORT
# ==========================

@app.route("/daily_report")
@login_required
def daily_report():
    selected_date = request.args.get(
        "date",
        datetime.now().strftime("%Y-%m-%d")
    )

    payments = Payment.query.filter_by(
        payment_date=selected_date,
        user_id=current_user.id
    ).all()

    total_collection = sum(payment.amount for payment in payments)

    return render_template(
        "daily_report.html",
        payments=payments,
        selected_date=selected_date,
        total_collection=total_collection
    )


# ==========================
# PENDING REPORT
# ==========================

@app.route("/pending_report")
@login_required
def pending_report():
    customers = Customer.query.filter(
        Customer.remaining_balance > 0,
        Customer.user_id == current_user.id
    ).all()

    total_pending = sum(
        customer.remaining_balance for customer in customers
    )

    return render_template(
        "pending_report.html",
        customers=customers,
        total_pending=total_pending
    )


# ==========================
# EXPORT DAILY REPORT EXCEL
# ==========================

@app.route("/export_daily_report_excel/<date>")
@login_required
def export_daily_report_excel(date):
    payments = Payment.query.filter_by(
        payment_date=date,
        user_id=current_user.id
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"
    ws.append(["Date", "Customer ID", "Amount"])

    for payment in payments:
        ws.append([payment.payment_date, payment.customer_id, payment.amount])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"Daily_Report_{date}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==========================
# EXPORT DAILY REPORT PDF
# ==========================

@app.route("/export_daily_report_pdf/<date>")
@login_required
def export_daily_report_pdf(date):
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4

    normal_font, bold_font, has_tamil = _register_tamil_font()

    payments = Payment.query.filter_by(
        payment_date=date,
        user_id=current_user.id
    ).all()

    total_collection = sum(p.amount for p in payments)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()

    _reg = _FONT_CACHE.get("result", (None, "Helvetica", "Helvetica-Bold", False))
    base_font = _reg[1]   # NotoSans  (Latin)
    hdr_font  = _reg[2]   # NotoSans-Bold

    title_style = ParagraphStyle(
        "DailyTitle",
        parent=styles["Title"],
        fontName=hdr_font
    )
    normal_style = ParagraphStyle(
        "DailyNormal",
        parent=styles["Normal"],
        fontName=base_font
    )

    elements = []
    elements.append(Paragraph(_pdf_text(f"Daily Report - {date}"), title_style))
    elements.append(Spacer(1, 12))

    data = [["No", "Customer ID", "Date", "Amount"]]
    for i, p in enumerate(payments, start=1):
        data.append([
            Paragraph(_pdf_text(str(i)), normal_style),
            Paragraph(_pdf_text(str(p.customer_id)), normal_style),
            Paragraph(_pdf_text(p.payment_date), normal_style),
            Paragraph(_pdf_text(f"RS-{p.amount}"), normal_style)
        ])

    data.append([
        "",
        Paragraph(_pdf_text("TOTAL"), normal_style),
        "",
        Paragraph(_pdf_text(f"RS-{total_collection}"), normal_style)
    ])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("BACKGROUND", (0, -1), (-1, -1), colors.lightyellow),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTNAME", (0, 0), (-1, 0), hdr_font),
        ("FONTNAME", (0, -1), (-1, -1), hdr_font),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Daily_Report_{date}.pdf",
        mimetype="application/pdf"
    )


# ==========================
# EXPORT MONTHLY COLLECTION EXCEL
# ==========================

@app.route("/export_collection/<month>")
@login_required
def export_collection(month):
    from sqlalchemy import func, cast, Integer
    from collections import defaultdict

    wb = Workbook()
    ws = wb.active
    ws.title = "Collection Sheet"

    headers = ["ID", "Name", "Loan"]
    for day in range(1, 32):
        headers.append(str(day))
    headers.extend(["Month Total", "Total Paid", "Balance", "Status"])
    ws.append(headers)

    customers = Customer.query.filter_by(user_id=current_user.id).all()
    customers = _sort_customers(customers)

    # ------------------------------------------------------------------
    # FIX (N+1 query / performance): the old code ran ONE
    # Payment.query.filter_by(...).all() PER CUSTOMER inside this loop,
    # plus a second full-table Payment.query.filter_by(...).all() below
    # for the summary row - for ~1,000 customers that's 1,000+ separate
    # round trips to Supabase on every export. This is the exact N+1
    # problem that was already fixed in export_collection_pdf but was
    # left unfixed here in the Excel export. We now pull every payment
    # for this user+month ONCE, pre-aggregated (SUM + GROUP BY) inside
    # Postgres, into a small in-memory lookup:
    #   payments_by_customer[customer_id][day] = total_amount_that_day
    # ------------------------------------------------------------------
    day_expr = cast(func.substr(Payment.payment_date, 9, 2), Integer)
    payment_query = (
        db.session.query(
            Payment.customer_id,
            day_expr.label("day"),
            func.sum(Payment.amount).label("day_total")
        )
        .filter(
            Payment.user_id == current_user.id,
            Payment.payment_date.like(f"{month}%")
        )
        .group_by(Payment.customer_id, day_expr)
    )

    payments_by_customer = defaultdict(dict)   # {customer_id: {day: amount}}
    day_totals = defaultdict(float)            # {day: total across ALL customers}
    month_total_all = 0

    for cust_id, day, day_total in payment_query.yield_per(500):
        payments_by_customer[cust_id][day] = day_total
        day_totals[day] += day_total
        month_total_all += day_total

    for customer in customers:
        row = [customer.customer_id, customer.name, customer.loan_amount]
        month_total = 0
        cust_payments = payments_by_customer.get(customer.customer_id, {})

        for day in range(1, 32):
            amt = cust_payments.get(day)
            if amt is None:
                row.append("-")
            else:
                row.append(amt)
                month_total += amt

        row.extend([
            month_total,
            customer.total_paid,
            customer.remaining_balance,
            customer.status
        ])
        ws.append(row)

    # DAY TOTAL SUMMARY ROW - built entirely from the aggregates computed
    # in the single pass above (no second full-table Payment query).
    total_paid_all = sum(c.total_paid for c in customers)
    total_balance_all = sum(c.remaining_balance for c in customers)

    summary_row = ["DAY TOTAL", "", ""]
    for day in range(1, 32):
        summary_row.append(day_totals.get(day, 0))

    summary_row.extend([month_total_all, total_paid_all, total_balance_all, "-"])
    ws.append(summary_row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"Monthly_Collection_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==========================
# EXPORT MONTHLY COLLECTION PDF
# (Tamil Language Rendering Fixed)
# ==========================

@app.route("/export_collection_pdf/<month>")
@login_required
def export_collection_pdf(month):
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    )
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import landscape, A3
    from sqlalchemy import func, cast, Integer
    from collections import defaultdict

    # Register Tamil-capable font (cached globally after first call)
    normal_font, bold_font, has_tamil = _register_tamil_font()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A3))
    styles = getSampleStyleSheet()

    _reg = _FONT_CACHE.get("result", (None, "Helvetica", "Helvetica-Bold", False))
    base_font = _reg[1]   # NotoSans  (Latin)
    hdr_font  = _reg[2]   # NotoSans-Bold

    title_style = ParagraphStyle(
        "MonthlyTitle",
        parent=styles["Title"],
        fontName=hdr_font
    )
    normal_style = ParagraphStyle(
        "MonthlyNormal",
        parent=styles["Normal"],
        fontName=base_font,
        fontSize=8
    )

    elements = []
    elements.append(
        Paragraph(_pdf_text(f"Monthly Collection Sheet - {month}"), title_style)
    )
    elements.append(Spacer(1, 10))

    headers = ["ID", "Name", "Loan"]
    for day in range(1, 32):
        headers.append(str(day))
    headers.extend(["Month Total", "Total Paid", "Balance", "Status"])

    # OPTIMIZATION 1 - reuse a single TableStyle object for every chunk
    # instead of rebuilding an identical TableStyle 100+ times.
    row_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ])

    # =====================================================================
    # OPTIMIZATION 2 - KILL THE N+1 QUERY
    # The old code ran one Payment.query.filter_by(...).all() PER CUSTOMER
    # (1000 customers = 1000+ round trips to Supabase - this alone is what
    # was blowing up connections / time on Render). We instead pull every
    # payment for this month ONCE, pre-aggregated (SUM + GROUP BY) inside
    # Postgres itself, and build a small in-memory lookup dict:
    #   payments_by_customer[customer_id][day] = total_amount_that_day
    # This dict is tiny (at most 31 entries per customer who paid) compared
    # to holding every raw Payment row in memory.
    # =====================================================================
    day_expr = cast(func.substr(Payment.payment_date, 9, 2), Integer)

    payment_query = (
        db.session.query(
            Payment.customer_id,
            day_expr.label("day"),
            func.sum(Payment.amount).label("day_total")
        )
        .filter(
            Payment.user_id == current_user.id,
            Payment.payment_date.like(f"{month}%")
        )
        .group_by(Payment.customer_id, day_expr)
    )

    payments_by_customer = defaultdict(dict)   # {customer_id: {day: amount}}
    day_totals = defaultdict(float)            # {day: total across ALL customers}
    month_total_all = 0

    # OPTIMIZATION 3 - stream the aggregated rows instead of .all().
    # yield_per() fetches in batches from the DB cursor rather than
    # materializing the entire result set in Python memory at once.
    for cust_id, day, day_total in payment_query.yield_per(500):
        payments_by_customer[cust_id][day] = day_total
        day_totals[day] += day_total
        month_total_all += day_total

    # =====================================================================
    # OPTIMIZATION 4 - NO sorted() ON QUERY RESULTS, LET POSTGRES ORDER BY
    # The old code did Customer.query...yield_per(100) and then immediately
    # threw away the streaming benefit by calling sorted() on it, which
    # forces the ENTIRE result set into a Python list before sorting.
    # For customer_id values that are numeric strings ("1", "2", ... "9999"),
    # ORDER BY LENGTH(customer_id), customer_id in Postgres reproduces the
    # exact same order as sorting by int(customer_id) in Python (shorter
    # numeric strings always sort first, then lexicographically). This lets
    # the database do the sort and lets us stream rows straight off the
    # cursor with constant memory, regardless of 100 vs 10,000 customers.
    # =====================================================================
    customers_query = (
        Customer.query
        .filter_by(user_id=current_user.id)
        .order_by(func.length(Customer.customer_id), Customer.customer_id)
        .yield_per(200)
    )

    def flush_chunk(rows, include_header):
        """
        OPTIMIZATION 5 - CHUNKED TABLES + PAGE BREAKS INSTEAD OF ONE
        GIANT TABLE.
        A single ReportLab Table() holding thousands of rows keeps every
        Paragraph/cell flowable alive in memory simultaneously while
        doc.build() lays out pages - this is the #1 cause of the SIGKILL
        / OOM crash on Render's free tier. Instead we build a small
        Table() per 100 rows, append it to `elements`, and let Python's
        GC reclaim the row list. Column layout, fonts, grid lines and
        colors are identical to the original design.
        """
        chunk_data = ([headers] + rows) if include_header else rows
        t = Table(chunk_data, repeatRows=1 if include_header else 0)
        t.setStyle(row_style)
        elements.append(t)

    data_rows = []
    row_count_in_chunk = 0
    is_first_chunk = True
    total_paid_all = 0
    total_balance_all = 0

    for customer in customers_query:
        # _pdf_text() switches font per-character: Tamil->NotoSansTamil, Latin->NotoSans
        name_cell = Paragraph(_pdf_text(customer.name), normal_style)
        row = [customer.customer_id, name_cell, customer.loan_amount]

        cust_payments = payments_by_customer.get(customer.customer_id, {})
        month_total = 0
        for day in range(1, 32):
            amt = cust_payments.get(day)
            if amt is None:
                row.append("-")
            else:
                row.append(amt)
                month_total += amt

        row.extend([
            month_total,
            customer.total_paid,
            customer.remaining_balance,
            customer.status
        ])
        data_rows.append(row)
        row_count_in_chunk += 1

        total_paid_all += customer.total_paid or 0
        total_balance_all += customer.remaining_balance or 0

        # OPTIMIZATION 6 - flush + PageBreak every 100 rows so no single
        # Table() or in-memory list ever holds more than 100 rows.
        if row_count_in_chunk == 100:
            flush_chunk(data_rows, include_header=is_first_chunk)
            elements.append(PageBreak())
            data_rows = []          # release the chunk, don't accumulate
            row_count_in_chunk = 0
            is_first_chunk = False

    # Flush the final partial chunk (< 100 rows)
    if data_rows:
        flush_chunk(data_rows, include_header=is_first_chunk)
        data_rows = []

    # OPTIMIZATION 7 - DAY TOTAL ROW built entirely from the aggregates
    # computed during the single streaming pass above (no second
    # Payment.query.filter_by(...).all() needed, unlike the original).
    total_row = ["DAY TOTAL", "", ""]
    for day in range(1, 32):
        total_row.append(day_totals.get(day, 0))
    total_row.extend([month_total_all, total_paid_all, total_balance_all, "-"])

    total_table = Table([total_row])
    total_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.lightgreen),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTNAME", (0, 0), (-1, -1), bold_font),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(total_table)

    # OPTIMIZATION 8 - release the SQLAlchemy identity map. After
    # streaming thousands of Customer objects, the Session still holds
    # references to all of them. expunge_all() detaches them so they can
    # be garbage-collected before/while doc.build() does its (memory
    # heavier) layout pass.
    db.session.expunge_all()

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Monthly_Collection_{month}.pdf",
        mimetype="application/pdf"
    )


# ==========================
# CUSTOMER LEDGER
# ==========================

@app.route("/customer_ledger")
@login_required
def customer_ledger():
    customers = Customer.query.filter_by(user_id=current_user.id).all()
    customers = _sort_customers(customers)

    # FIX (Bug): customer_ledger.html reads customer.alert to render the
    # Alert column and drive the "Collection Required" filter, but that
    # attribute was never set here -> every row silently showed "Active"
    # and the filter option matched nothing.
    for customer in customers:
        customer.alert = _compute_alert(customer)

    return render_template("customer_ledger.html", customers=customers)


# ==========================
# EXPORT CUSTOMER LEDGER EXCEL
# (Fixed: now filters by user_id)
# ==========================

@app.route("/export_customers")
@login_required
def export_customers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = [
        "Customer ID", "Name", "Mobile", "Loan Amount",
        "Daily Due", "Total Paid", "Balance", "Status"
    ]
    ws.append(headers)

    # FIX: filter by current user (was missing user_id filter)
    customers = Customer.query.filter_by(user_id=current_user.id).all()
    customers = _sort_customers(customers)

    for customer in customers:
        ws.append([
            customer.customer_id,
            customer.name,
            customer.mobile,
            customer.loan_amount,
            customer.daily_due,
            customer.total_paid,
            customer.remaining_balance,
            customer.status
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="customer_ledger.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==========================
# CUSTOMER DETAILS
# ==========================

@app.route("/customer/<customer_id>")
@login_required
def customer_details(customer_id):
    customer = Customer.query.filter_by(
        customer_id=customer_id,
        user_id=current_user.id
    ).first_or_404()

    payments = Payment.query.filter_by(
        customer_id=customer_id,
        user_id=current_user.id
    ).all()

    # Calculate day-based fields from start_date and end_date
    total_days = "-"
    days_passed = "-"
    remaining_days = "-"

    try:
        if customer.start_date and customer.end_date:
            fmt = "%Y-%m-%d"
            start = datetime.strptime(str(customer.start_date), fmt)
            end   = datetime.strptime(str(customer.end_date),   fmt)
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

            # Inclusive counting: the start date itself counts as Day 1.
            total_days = (end - start).days + 1

            if today < start:
                # Loan hasn't started yet
                days_passed = 0
            else:
                days_passed = (today - start).days + 1
                days_passed = min(days_passed, total_days)

            remaining_days = max(0, total_days - days_passed)
    except (ValueError, TypeError):
        pass  # leave defaults as "-" if dates are missing / malformed

    return render_template(
        "customer_details.html",
        customer=customer,
        payments=payments,
        total_days=total_days,
        days_passed=days_passed,
        remaining_days=remaining_days
    )


# ==========================
# EDIT CUSTOMER
# ==========================

@app.route("/edit_customer/<customer_id>", methods=["GET", "POST"])
@login_required
def edit_customer(customer_id):
    customer = Customer.query.filter_by(
        customer_id=customer_id,
        user_id=current_user.id
    ).first_or_404()

    if request.method == "POST":
        # FIX: float() on bad/missing input used to raise an uncaught
        # ValueError -> 500 error page instead of a friendly flash message.
        try:
            new_loan_amount = float(request.form["loan_amount"])
            new_daily_due = float(request.form["daily_due"])
        except (ValueError, TypeError):
            flash("Loan Amount and Daily Due must be valid numbers.", "danger")
            return redirect(url_for("edit_customer", customer_id=customer_id))

        if new_loan_amount < 0 or new_daily_due < 0:
            flash("Loan Amount and Daily Due cannot be negative.", "danger")
            return redirect(url_for("edit_customer", customer_id=customer_id))

        customer.name        = request.form["name"]
        customer.mobile      = request.form["mobile"]
        customer.address     = request.form.get("address", customer.address or "")
        customer.loan_amount = new_loan_amount
        customer.daily_due   = new_daily_due
        customer.end_date    = request.form.get("end_date", customer.end_date or "")
        customer.remaining_balance = customer.loan_amount - customer.total_paid

        # Re-apply the same Closed/Active rule used everywhere else so
        # editing the loan amount can't leave status out of sync with
        # the recalculated balance (e.g. increasing the loan on a
        # previously "Closed" customer used to leave it stuck "Closed").
        if customer.remaining_balance <= 0:
            customer.remaining_balance = 0
            customer.status = "Closed"
        else:
            customer.status = "Active"

        # FIX (Missing rollback): a DB error mid-commit (lost connection,
        # constraint violation) previously left the session in a broken
        # state with no rollback, which can poison subsequent requests
        # sharing the same pooled connection.
        try:
            db.session.commit()
            flash("Customer Updated Successfully")
        except Exception as e:
            db.session.rollback()
            flash(f"Error updating customer: {str(e)}", "danger")
        return redirect(url_for("customer_ledger"))

    return render_template("edit_customer.html", customer=customer)


# ==========================
# DELETE CUSTOMER
# ==========================

@app.route("/delete_customer/<customer_id>")
@login_required
def delete_customer(customer_id):
    customer = Customer.query.filter_by(
        customer_id=customer_id,
        user_id=current_user.id
    ).first()

    if customer:
        Payment.query.filter_by(
            customer_id=customer_id,
            user_id=current_user.id
        ).delete()

        db.session.delete(customer)
        db.session.commit()
        flash("Customer Deleted Successfully")

    return redirect(url_for("customer_ledger"))


# ==========================
# EDIT PAYMENT
# ==========================

@app.route("/edit_payment/<int:payment_id>", methods=["GET", "POST"])
@login_required
def edit_payment(payment_id):
    payment = Payment.query.filter_by(
        id=payment_id,
        user_id=current_user.id
    ).first_or_404()

    if request.method == "POST":
        # FIX (Transaction safety): lock the customer row, same reasoning
        # as add_payment - prevents lost updates if two edits land
        # concurrently on the same customer's balance.
        customer = Customer.query.filter_by(
            customer_id=payment.customer_id,
            user_id=current_user.id
        ).with_for_update().first()

        # FIX (Bug): if the customer record this payment points to was
        # deleted separately (data drift) this used to raise
        # AttributeError on customer.total_paid below -> 500 error.
        if not customer:
            flash("Cannot update this payment: its customer record no longer exists.", "danger")
            return redirect(url_for("dashboard"))

        old_amount = payment.amount
        try:
            new_amount = float(request.form["amount"])
        except (ValueError, TypeError):
            flash("Payment amount must be a valid number.", "danger")
            return redirect(url_for("edit_payment", payment_id=payment_id))

        if new_amount < 0:
            flash("Payment amount cannot be negative.", "danger")
            return redirect(url_for("edit_payment", payment_id=payment_id))

        new_date = request.form["payment_date"]

        # If moving this payment to a date that already has a different
        # payment for the same customer, merge into that one instead of
        # creating a second row for the same day.
        duplicate = Payment.query.filter(
            Payment.customer_id == payment.customer_id,
            Payment.payment_date == new_date,
            Payment.user_id == current_user.id,
            Payment.id != payment.id
        ).first()

        if duplicate:
            duplicate.amount += new_amount
            customer.total_paid = customer.total_paid - old_amount + new_amount
            db.session.delete(payment)
            flash("Merged into existing payment for that date")
        else:
            payment.payment_date = new_date
            payment.amount = new_amount
            customer.total_paid = customer.total_paid - old_amount + new_amount
            flash("Payment Updated Successfully")

        customer.remaining_balance = customer.loan_amount - customer.total_paid

        if customer.remaining_balance <= 0:
            customer.remaining_balance = 0
            customer.status = "Closed"
        else:
            customer.status = "Active"

        db.session.commit()
        return redirect(url_for("customer_details", customer_id=customer.customer_id))

    return render_template("edit_payment.html", payment=payment)


# ==========================
# DELETE PAYMENT
# ==========================

@app.route("/delete_payment/<int:payment_id>")
@login_required
def delete_payment(payment_id):
    payment = Payment.query.filter_by(
        id=payment_id,
        user_id=current_user.id
    ).first_or_404()

    # FIX (Transaction safety): lock the row, same reasoning as add_payment.
    customer = Customer.query.filter_by(
        customer_id=payment.customer_id,
        user_id=current_user.id
    ).with_for_update().first()

    # FIX (Bug): previously this crashed with AttributeError (500) if the
    # customer record no longer existed for this payment.
    if not customer:
        db.session.delete(payment)
        db.session.commit()
        flash("Payment Deleted Successfully (its customer record was already gone)")
        return redirect(url_for("dashboard"))

    customer.total_paid -= payment.amount
    customer.remaining_balance = customer.loan_amount - customer.total_paid

    if customer.remaining_balance <= 0:
        customer.remaining_balance = 0
        customer.status = "Closed"
    else:
        customer.status = "Active"

    db.session.delete(payment)
    db.session.commit()

    flash("Payment Deleted Successfully")
    return redirect(url_for("customer_details", customer_id=customer.customer_id))


# ==========================
# EXPORT SINGLE CUSTOMER EXCEL
# ==========================

@app.route("/export_customer/<customer_id>")
@login_required
def export_customer(customer_id):
    customer = Customer.query.filter_by(
        customer_id=customer_id,
        user_id=current_user.id
    ).first_or_404()

    payments = Payment.query.filter_by(
        customer_id=customer_id,
        user_id=current_user.id
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Statement"

    ws.append(["Customer ID", customer.customer_id])
    ws.append(["Name", customer.name])
    ws.append(["Mobile", customer.mobile])
    ws.append(["Loan Amount", customer.loan_amount])
    ws.append(["Daily Due", customer.daily_due])
    ws.append(["Total Paid", customer.total_paid])
    ws.append(["Balance", customer.remaining_balance])
    ws.append(["Status", customer.status])
    ws.append([])
    ws.append(["Payment History"])
    ws.append(["Date", "Amount"])

    for payment in payments:
        ws.append([payment.payment_date, payment.amount])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"{customer.customer_id}_statement.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==========================
# CUSTOMER STATEMENT PDF
# (Tamil Language Rendering Fixed)
# ==========================

@app.route("/customer_statement_pdf/<customer_id>")
@login_required
def customer_statement_pdf(customer_id):
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4

    customer = Customer.query.filter_by(
        customer_id=customer_id,
        user_id=current_user.id
    ).first_or_404()

    payments = Payment.query.filter_by(
        customer_id=customer_id,
        user_id=current_user.id
    ).all()

    # Register both Tamil and Latin fonts
    normal_font, bold_font, has_tamil = _register_tamil_font()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()

    # Use NotoSans (Latin) as base font; _pdf_text() switches to NotoSansTamil
    # per-character for Tamil Unicode ranges, so both scripts render correctly.
    # _register_tamil_font returns (tamil_font, bold_font, has_tamil);
    # retrieve the latin font name directly from the cache for the base style.
    _reg = _FONT_CACHE.get("result", (normal_font, normal_font, bold_font, has_tamil))
    base_font = _reg[1]   # latin_font (NotoSans)
    hdr_font  = _reg[2]   # bold font  (NotoSans-Bold)

    title_style = ParagraphStyle(
        "StmtTitle",
        parent=styles["Title"],
        fontName=hdr_font
    )
    normal_style = ParagraphStyle(
        "StmtNormal",
        parent=styles["Normal"],
        fontName=base_font
    )

    elements = []
    elements.append(Paragraph(_pdf_text("CUSTOMER STATEMENT"), title_style))
    elements.append(Spacer(1, 12))

    # _pdf_text() wraps Tamil chars with NotoSansTamil, Latin stays NotoSans
    info_lines = [
        _pdf_text(f"Customer ID: {customer.customer_id}"),
        _pdf_text(f"Name: {customer.name}"),
        _pdf_text(f"Mobile: {customer.mobile}"),
        _pdf_text(f"Loan Amount: RS-{customer.loan_amount}"),
        _pdf_text(f"Total Paid: RS-{customer.total_paid}"),
        _pdf_text(f"Balance: RS-{customer.remaining_balance}"),
    ]
    for line in info_lines:
        elements.append(Paragraph(line, normal_style))

    elements.append(Spacer(1, 15))

    data = [["No", "Date", "Amount"]]
    for index, payment in enumerate(payments, start=1):
        data.append([
            Paragraph(_pdf_text(str(index)), normal_style),
            Paragraph(_pdf_text(payment.payment_date), normal_style),
            Paragraph(_pdf_text(f"RS-{payment.amount}"), normal_style)
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONTNAME", (0, 0), (-1, 0), hdr_font),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{customer.customer_id}_statement.pdf",
        mimetype="application/pdf"
    )


# ==========================
# BACKUP DATABASE
# ==========================

@app.route("/backup_database")
@login_required
def backup_database():
    # NOTE: the app now runs on PostgreSQL (Supabase), so there is no local
    # .db file to send anymore. Export all of this user's data to Excel
    # instead, with one sheet per table, so backups actually work.
    wb = Workbook()

    ws_customers = wb.active
    ws_customers.title = "Customers"
    ws_customers.append([
        "Customer ID", "Name", "Mobile", "Address", "Loan Amount",
        "Daily Due", "Total Paid", "Remaining Balance", "Status",
        "Start Date", "End Date"
    ])
    customers = Customer.query.filter_by(user_id=current_user.id).all()
    for c in customers:
        ws_customers.append([
            c.customer_id, c.name, c.mobile, c.address, c.loan_amount,
            c.daily_due, c.total_paid, c.remaining_balance, c.status,
            c.start_date, c.end_date
        ])

    ws_payments = wb.create_sheet("Payments")
    ws_payments.append(["Customer ID", "Payment Date", "Amount"])
    payments = Payment.query.filter_by(user_id=current_user.id).all()
    for p in payments:
        ws_payments.append([p.customer_id, p.payment_date, p.amount])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    backup_name = f"finance_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=backup_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==========================
# RESTORE DATABASE
# ==========================

@app.route("/restore_database", methods=["GET", "POST"])
@login_required
def restore_database():
    if request.method == "POST":
        backup_file = request.files.get("backup_file")

        if backup_file and backup_file.filename.endswith(".xlsx"):
            from openpyxl import load_workbook
            try:
                wb = load_workbook(backup_file, data_only=True)

                # Wipe this user's existing data before restoring
                Payment.query.filter_by(user_id=current_user.id).delete()
                Customer.query.filter_by(user_id=current_user.id).delete()

                if "Customers" in wb.sheetnames:
                    ws = wb["Customers"]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or row[0] is None:
                            continue
                        (customer_id, name, mobile, address, loan_amount,
                         daily_due, total_paid, remaining_balance, status,
                         start_date, end_date) = row
                        db.session.add(Customer(
                            customer_id=str(customer_id),
                            name=name,
                            mobile=mobile,
                            address=address,
                            loan_amount=loan_amount or 0,
                            daily_due=daily_due or 0,
                            total_paid=total_paid or 0,
                            remaining_balance=remaining_balance or 0,
                            status=status or "Active",
                            start_date=start_date,
                            end_date=end_date,
                            user_id=current_user.id
                        ))

                if "Payments" in wb.sheetnames:
                    ws = wb["Payments"]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or row[0] is None:
                            continue
                        customer_id, payment_date, amount = row
                        db.session.add(Payment(
                            customer_id=str(customer_id),
                            payment_date=str(payment_date),
                            amount=amount or 0,
                            user_id=current_user.id
                        ))

                db.session.commit()
                flash("Database Restored Successfully")
            except Exception as e:
                db.session.rollback()
                flash(f"Restore failed: {str(e)}", "danger")

            return redirect(url_for("dashboard"))

        flash("Please upload a valid .xlsx backup file")

    return render_template("restore_database.html")


# ==========================
# COMPANY SETTINGS
# ==========================

@app.route("/company_settings", methods=["GET", "POST"])
@login_required
def company_settings():
    settings = CompanySettings.query.filter_by(user_id=current_user.id).first()

    if not settings:
        settings = CompanySettings(user_id=current_user.id)
        db.session.add(settings)
        db.session.commit()

    if request.method == "POST":
        settings.company_name = request.form["company_name"]
        settings.address = request.form["address"]
        settings.phone = request.form["phone"]

        try:
            db.session.commit()
            flash("Settings Saved Successfully")
        except Exception as e:
            db.session.rollback()
            flash(f"Error saving settings: {str(e)}", "danger")
        return redirect(url_for("company_settings"))

    return render_template("company_settings.html", settings=settings)


# ==========================
# LOGOUT
# ==========================

@app.route("/logout")
@login_required
def logout():
    logout_user()
    session.clear()
    return redirect(url_for("login"))

# ==========================
# DATABASE INIT
# ==========================

with app.app_context():
    db.create_all()

    admin = Admin.query.filter_by(username="admin").first()

    if not admin:
        admin = Admin(
            username="admin",
            mobile="9999999999",
            password=generate_password_hash("admin123")
        )

        db.session.add(admin)
        try:
            db.session.commit()
            logger.info("Default admin account created (username: admin).")
        except IntegrityError:
            # FIX (Deployment): this module runs once per gunicorn worker
            # process on Render. With more than one worker, two workers
            # can both see "no admin yet" and both try to insert the same
            # row at startup; the loser previously crashed the whole
            # worker with an unhandled IntegrityError instead of just
            # noticing the other worker already created it.
            db.session.rollback()


        # ==========================
        # RUN APP
        # ==========================

if __name__ == "__main__":
    app.run(
        host="127.0.0.1",
        port=5000,
        debug=False,
        use_reloader=False
    )